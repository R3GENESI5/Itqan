import re, io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

SRC = "KawakibNayyirat.Shamela0000309.mARkdown"
raw = io.open(SRC, encoding="utf-8").read().split("\n")

# --- strip metadata header, capture a few fields ---
start = next(i for i,l in enumerate(raw) if l.strip()=="#META#Header#End#") + 1
meta = {}
for l in raw[:start]:
    m = re.match(r"#META#\s*\d+\.(\w+)\s*::\s*(.*)", l)
    if m and m.group(2).strip() not in ("NODATA","NOTGIVEN",""):
        meta.setdefault(m.group(1), m.group(2).strip())
body = raw[start:]

# --- merge soft-wrap continuation lines ---
logical = []
for l in body:
    if l.startswith("~~"):
        logical[-1] += " " + l[2:]
    else:
        logical.append(l)

PAGE = re.compile(r"PageV(\d+)P(\d+)")
def pages_in(text):
    out = []
    for v,p in PAGE.findall(text):
        v,p = int(v), int(p)
        if not (v==0 and p==0):
            out.append(f"ج{v}/ص{p}")
    return out
def strip_tokens(text):
    text = PAGE.sub("", text)
    text = re.sub(r"(?<!\S)ms\d+(?!\S)", "", text)
    text = re.sub(r"[ \t]{2,}", " ", text)
    return text.strip()

# --- walk: track section (### |), collect entries (### $) ---
records = []
muqaddima = []
section = ""
cur = None
def flush():
    global cur
    if cur:
        cur["pages"] = pages_in(cur["_raw"])
        cur["text"] = strip_tokens(cur["_raw"])
        del cur["_raw"]
        records.append(cur)
        cur = None

in_front = True
for l in logical:
    s = l.strip()
    if not s or s.startswith("######OpenITI"):
        continue
    if s.startswith("### |"):
        flush()
        title = re.sub(r"^###\s*\|\s*", "", s).strip()
        if title.upper().startswith("APPENDIX"):
            continue
        section = title
        in_front = (section == "مقدمة المصنف")
        continue
    if s.startswith("### $"):
        flush()
        in_front = False
        h = re.sub(r"^###\s*\$+\s*", "", s).strip()
        m = re.match(r"(\d+)\s*[-–]\s*(.*)", h)
        num = m.group(1) if m else ""
        name = (m.group(2) if m else h).strip()
        cur = {"section": section, "num": num, "name": strip_tokens(name), "_raw": ""}
        continue
    # plain content line
    s = re.sub(r"^#{1,6}\s*", "", s)
    if cur is not None:
        cur["_raw"] += " " + s
    elif in_front:
        muqaddima.append(s)
flush()

# ---------------- build workbook ----------------
wb = Workbook()
thin = Side(style="thin", color="DDDDDD")
border = Border(left=thin,right=thin,top=thin,bottom=thin)
hdr_fill = PatternFill("solid", fgColor="1F4E78")
hdr_font = Font(bold=True, color="FFFFFF", size=12)
arab = Alignment(horizontal="right", vertical="top", wrap_text=True, readingOrder=2)
center = Alignment(horizontal="center", vertical="top")

# Sheet 1: narrators
ws = wb.active
ws.title = "المختلطون"
ws.sheet_view.rightToLeft = True
cols = [("#", 6), ("الباب / القسم", 22), ("الراوي", 40),
        ("الصفحة", 12), ("الصفحات", 18), ("النص (أقوال النقاد)", 110), ("أحرف", 8)]
ws.append([c[0] for c in cols])
for i,(t,w) in enumerate(cols, 1):
    cell = ws.cell(row=1, column=i)
    cell.fill = hdr_fill; cell.font = hdr_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"
ws.row_dimensions[1].height = 26

for idx, r in enumerate(records, start=1):
    pages = r["pages"]
    row = [idx, r["section"], r["name"],
           pages[0] if pages else "",
           "، ".join(pages),
           r["text"], len(r["text"])]
    ws.append(row)
    rr = ws.max_row
    for c in range(1, len(cols)+1):
        cell = ws.cell(row=rr, column=c)
        cell.border = border
        cell.alignment = center if c in (1,4,7) else arab
    ws.row_dimensions[rr].height = 90

# Sheet 2: muqaddima
ws2 = wb.create_sheet("مقدمة المصنف")
ws2.sheet_view.rightToLeft = True
ws2.column_dimensions["A"].width = 140
ws2.cell(row=1, column=1, value="مقدمة المصنف").font = Font(bold=True, size=13)
muq = strip_tokens(" ".join(muqaddima))
ws2.cell(row=2, column=1, value=muq).alignment = arab
ws2.row_dimensions[2].height = 600

# Sheet 3: about / provenance
ws3 = wb.create_sheet("معلومات")
ws3.sheet_view.rightToLeft = True
ws3.column_dimensions["A"].width = 26; ws3.column_dimensions["B"].width = 90
info = [
 ("الكتاب", "الكواكب النيرات في معرفة من اختلط من الرواة الثقات"),
 ("المؤلف", meta.get("AuthorNAME","أبو البركات محمد بن أحمد بن الكيال (ت 939هـ)")),
 ("المحقق", meta.get("EdEDITOR","عبد القيوم عبد رب النبي")),
 ("الناشر", f'{meta.get("EdPUBLISHER","دار المأمون ـ بيروت")} — {meta.get("EdNUMBER","الأولى 1981م")}'),
 ("المصدر الرقمي", "OpenITI corpus — 0939IbnKayyal.KawakibNayyirat (Shamela0000309, النسخة الأساسية المنقّحة)"),
 ("عدد التراجم", str(len(records))),
 ("الأقسام", str(len({r["section"] for r in records}))),
 ("ملاحظة", "النص منقول حرفيًا من النسخة الأساسية؛ روابط الصفحات بصيغة ج/ص. لا استنتاج آلي للأحكام."),
]
for i,(k,v) in enumerate(info, start=1):
    ws3.cell(row=i, column=1, value=k).font = Font(bold=True)
    ws3.cell(row=i, column=1).alignment = arab
    ws3.cell(row=i, column=2, value=v).alignment = arab

OUT = "al-Kawakib-al-Nayyirat.xlsx"
wb.save(OUT)
print("rows:", len(records), "| sections:", len({r['section'] for r in records}))
print("saved:", OUT)
# quick peek
for r in records[:3]:
    print(f"  #{r['num']:>3} [{r['section']}] {r['name'][:40]} | {r['pages'][:2]} | {len(r['text'])} chars")
