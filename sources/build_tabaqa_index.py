#!/usr/bin/env python3
"""Assign a generation (tabaqa) to each unified narrator and arrange by it,
so same-generation entries cluster and cross-book duplicates are easy to spot.

Signals (priority): arsanad explicit tabaqa > in-text companion/Taqrib ordinal >
arsanad death-year band > arsanad head-match > in-text tabi'/atba' > unknown.
A `tabaqa_basis` column records how each label was derived (auditable)."""
import csv, sys, re, os
from collections import defaultdict, Counter
csv.field_size_limit(sys.maxsize)
ROOT="/home/user/hadith"; OUT=os.path.join(ROOT,"sources/unified")

DIAC=re.compile("[ؐ-ؚـً-ٰٟۖ-ۭ]")
def norm(s):
    s=DIAC.sub("",s or "")
    s=(s.replace("أ","ا").replace("إ","ا").replace("آ","ا").replace("ٱ","ا")
        .replace("ى","ي").replace("ئ","ي").replace("ؤ","و").replace("ة","ه").replace("ء",""))
    s=re.sub(r"[^؀-ۿ\s]"," ",s)
    return re.sub(r"\s+"," ",s).strip()
def head(s,n=5): return " ".join(s.split()[:n])

# ---- ordinal word -> generation bin ----
ORD2BIN={"صحابي":1,"صحابيه":1,"الاولي":1,"الثانيه":2,"الثالثه":3,"الرابعه":3,
         "الخامسه":4,"السادسه":4,"السابعه":5,"الثامنه":5,"التاسعه":6,
         "العاشره":7,"الحاديه عشره":7,"الثانيه عشره":7}
ORD_RE=re.compile("("+"|".join(sorted(ORD2BIN,key=len,reverse=True))+")")
def ordinal_bin(s):
    m=ORD_RE.search(norm(s));  return ORD2BIN[m.group(1)] if m else None
def death_bin(s):
    m=re.search(r"\d+", s or "")
    if not m: return None
    d=int(m.group())
    return 2 if d<=110 else 3 if d<=150 else 4 if d<=180 else 5 if d<=215 else 6 if d<=245 else 7

# ---- Arabic numeral-word death-year parser (for "مات سنة ...") ----
_U={"احدي":1,"واحده":1,"واحد":1,"اثنتين":2,"اثنين":2,"ثنتين":2,"ثلاث":3,"ثلاثه":3,
    "اربع":4,"اربعه":4,"خمس":5,"خمسه":5,"ست":6,"سته":6,"سبع":7,"سبعه":7,
    "ثمان":8,"ثماني":8,"ثمانيه":8,"تسع":9,"تسعه":9}
_T={"عشرين":20,"ثلاثين":30,"اربعين":40,"خمسين":50,"ستين":60,"سبعين":70,"ثمانين":80,"تسعين":90,"عشره":10,"عشر":10}
_H={"مايه":100,"ميه":100,"مايتين":200,"مئتين":200,"مايتي":200,"ثلاثمايه":300,"اربعمايه":400,
    "خمسمايه":500,"ستمايه":600,"سبعمايه":700,"ثمانمايه":800,"تسعمايه":900}
YEAR_AT=re.compile(r"(?:مات|توفي|توفيت|ماتت)\s+(?:سنه\s+)?")
def parse_year(t):
    for m in YEAR_AT.finditer(t):
        toks=t[m.end():].split()[:8]; tot=0; i=0; got=False
        while i<len(toks):
            w=toks[i]
            if w in _H: tot+=_H[w]; got=True
            elif w in _T: tot+=_T[w]; got=True
            elif w in _U:
                if i+1<len(toks) and toks[i+1] in ("عشره","عشر"): tot+=_U[w]+10; i+=1
                else: tot+=_U[w]
                got=True
            elif w=="و": pass
            else: break
            i+=1
        if got and 1<=tot<=1000: return tot
    return None

# ---- in-text regexes (on normalized text) ----
COMP=re.compile(r"له صحبه|(?<![وا])صحابي\b|صحابي جليل|شهد بدرا|شهد احدا|بدري|هاجر الي")
TAQ =re.compile(r"من\s+(الاولي|الثانيه عشره|الحاديه عشره|الثانيه|الثالثه|الرابعه|الخامسه|السادسه|السابعه|الثامنه|التاسعه|العاشره)")
ATBA=re.compile(r"صغار اتباع التابعين|اتباع التابعين|تبع الاتباع")
TABI=re.compile(r"كبار التابعين|صغار التابعين|التابعين|تابعي")

# ---- arsanad lookup ----
ars=list(csv.DictReader(open(os.path.join(ROOT,"src/arsanad_narrators.csv"),encoding="utf-8")))
A_name={}; A_head={}
for r in ars:
    nn=norm(r["name"])
    if nn:
        A_name.setdefault(nn,r); A_head.setdefault(head(nn),r)
    sh=norm(r["shuhra"])
    if sh and r["shuhra"].strip()!="-": A_name.setdefault(sh,r)
def ars_bin(rec):
    b=ordinal_bin(rec["tabaqa"]);  basis="tabaqa"
    if b is None: b=death_bin(rec["death_year"]); basis="death"
    return b, basis

# ---- aggregate long-form per name_norm ----
agg=defaultdict(lambda:{"books":set(),"n":0,"types":set(),"pages":[],"disp":"",
                        "comp":False,"taq":None,"tabi":None,"dyr":None})
for r in csv.DictReader(open(os.path.join(OUT,"unified_narrator_index.csv"),encoding="utf-8-sig")):
    if r["keyable"]!="True": continue
    a=agg[r["name_norm"]]
    a["books"].add(r["source_slug"]); a["n"]+=1; a["types"].add(r["source_type"])
    if r["page"]: a["pages"].append(f'{r["source_slug"]}:{r["page"]}')
    if len(r["narrator_name"])>len(a["disp"]): a["disp"]=r["narrator_name"]
    t=norm(r["text"])
    if not a["comp"] and COMP.search(t): a["comp"]=True
    if a["taq"] is None:
        m=TAQ.search(t)
        if m: a["taq"]=ORD2BIN.get(m.group(1))
    if a["tabi"] is None:
        if ATBA.search(t): a["tabi"]=6 if "صغار" in t else 5
        elif TABI.search(t): a["tabi"]=2 if "كبار" in t else 4 if "صغار" in t else 3
    if a["dyr"] is None:
        y=parse_year(t)
        if y: a["dyr"]=y

BIN={1:"١ الصحابة",2:"٢ كبار التابعين",3:"٣ التابعون",4:"٤ صغار التابعين",
     5:"٥ أتباع التابعين",6:"٦ صغار أتباع التابعين",7:"٧ تبع الأتباع فمن بعدهم",8:"٨ غير محدد"}

def classify(name_norm, a):
    rec=A_name.get(name_norm)
    if rec:
        b,bs=ars_bin(rec)
        if b: return b, f"arsanad:{bs}"
    if a["taq"]:  return a["taq"],"text:طبقة"
    if a["dyr"]:
        db=death_bin(str(a["dyr"]))
        if db: return db,"text:وفاة"
    if a["comp"]: return 1,"text:صحابي"
    rec=A_head.get(head(name_norm))
    if rec:
        b,bs=ars_bin(rec)
        if b: return b, f"arsanad~:{bs}"
    if a["tabi"]: return a["tabi"],"text:تابعي"
    return 8,""

rows=[]
for k,a in agg.items():
    b,basis=classify(k,a)
    rows.append({"tabaqa_order":b,"tabaqa":BIN[b],"tabaqa_basis":basis,
                 "name_norm":k,"display_name":a["disp"],"n_books":len(a["books"]),
                 "n_entries":a["n"],"source_types":"; ".join(sorted(a["types"])),
                 "books":"; ".join(sorted(a["books"])),"pages":" | ".join(a["pages"])})
rows.sort(key=lambda r:(r["tabaqa_order"], r["name_norm"]))

GCOLS=["tabaqa_order","tabaqa","tabaqa_basis","name_norm","display_name","n_books","n_entries","source_types","books","pages"]
def dump(fn, rs):
    with open(fn,"w",encoding="utf-8-sig",newline="") as f:
        w=csv.writer(f); w.writerow(GCOLS)
        for r in rs: w.writerow([r[c] for c in GCOLS])
dump(os.path.join(OUT,"unified_by_tabaqa.csv"), rows)
cand=[r for r in rows if r["n_books"]>=2]
dump(os.path.join(OUT,"multi_source_candidates_by_tabaqa.csv"), cand)

# ---- xlsx: candidates, one sheet per generation, name-sorted, multi-book highlighted ----
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
wb=Workbook(); first=True
hf=PatternFill("solid",fgColor="1F4E78"); hfont=Font(bold=True,color="FFFFFF")
hi=PatternFill("solid",fgColor="FFF2CC")  # highlight >=3 books
arab=Alignment(horizontal="right",vertical="top",wrap_text=True,readingOrder=2)
ctr=Alignment(horizontal="center",vertical="top")
cols=[("الاسم",46),("# كتب",7),("# مواضع",8),("المصدر/الطبقة",18),("نوع المصدر",24),("الكتب",40),("الصفحات",38)]
for b in range(1,9):
    sub=[r for r in cand if r["tabaqa_order"]==b]
    if not sub: continue
    ws=(wb.active if first else wb.create_sheet()); first=False
    ws.title=BIN[b][:31]; ws.sheet_view.rightToLeft=True
    ws.append([c[0] for c in cols])
    for i,(t,wd) in enumerate(cols,1):
        c=ws.cell(1,i); c.fill=hf; c.font=hfont; c.alignment=Alignment(horizontal="center",wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width=wd
    ws.freeze_panes="A2"
    for r in sub:
        ws.append([r["display_name"],r["n_books"],r["n_entries"],r["tabaqa_basis"],
                   r["source_types"],r["books"],r["pages"][:500]])
        rr=ws.max_row
        for c in (1,4,5,6,7): ws.cell(rr,c).alignment=arab
        for c in (2,3): ws.cell(rr,c).alignment=ctr
        if r["n_books"]>=3:
            for c in range(1,8): ws.cell(rr,c).fill=hi
wb.save(os.path.join(OUT,"candidates_by_tabaqa.xlsx"))

# ---- report ----
dist=Counter(r["tabaqa"] for r in rows)
cdist=Counter(r["tabaqa"] for r in cand)
known=sum(1 for r in rows if r["tabaqa_order"]!=8)
ckn=sum(1 for r in cand if r["tabaqa_order"]!=8)
print(f"all narrators {len(rows)}: classified {known} ({100*known/len(rows):.0f}%)")
print(f"candidates(>=2) {len(cand)}: classified {ckn} ({100*ckn/len(cand):.0f}%)")
print("candidate distribution by generation:")
for b in range(1,9):
    if BIN[b] in cdist: print(f"   {BIN[b]:28} {cdist[BIN[b]]}")
print("basis breakdown (candidates):", dict(Counter(r["tabaqa_basis"] for r in cand if r["tabaqa_basis"])))
