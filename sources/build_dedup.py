#!/usr/bin/env python3
"""Conservative entity-resolution dedup for the unified narrator corpus.

Strategy (no naive name-merging):
  1. Block by name-core (ism + father) to keep it tractable + avoid cross-merges.
  2. Within a block, union-find merge two entries ONLY when the evidence is strong:
       - same canon canonical id (authoritative), OR
       - one full-nasab name contains the other (shorter >=4 tokens) AND death years
         don't conflict, OR
       - both death years known and within +/-2 AND names share their core.
  3. HARD SPLITS: entries linked to *different* canon ids are never merged;
     death years differing by >5 are never merged.
  4. Anchor each cluster to an canon id (=> canonical name/death/tabaqa) when any
     member matches.
Every merge records its evidence. Conservative by design: when unsure, keep apart.
"""
import csv, sys, re, ast, os
from collections import defaultdict, Counter
csv.field_size_limit(sys.maxsize)
ROOT="/home/user/hadith"; OUT=os.path.join(ROOT,"sources/unified")

DIAC=re.compile("[ؐ-ؚـً-ٰٟۖ-ۭ]")
def norm(s):
    s=DIAC.sub("",s or "")
    s=(s.replace("أ","ا").replace("إ","ا").replace("آ","ا").replace("ٱ","ا")
        .replace("ى","ي").replace("ئ","ي").replace("ؤ","و").replace("ة","ه").replace("ء",""))
    s=re.sub(r"[^؀-ۿ\s]"," ",s); return re.sub(r"\s+"," ",s).strip()
STOP={"بن","ابن","بنت","ابو","ابي","ابا","ام","عبد","ال","مولي","مولاه"}  # weak tokens for blocking
def content_toks(nn): return [t for t in nn.split() if t not in ("بن","ابن","بنت")]
# ubiquitous name elements — sharing only these is NOT identifying (prevents hub over-merge)
COMMON={"عبد","الله","الرحمن","الرحيم","محمد","احمد","علي","الحسن","الحسين","عمر","عثمان",
        "بكر","ابراهيم","اسماعيل","يحيي","عبيد","العزيز","الملك","السلام","الحميد","يوسف",
        "ابو","ابي","ابا","ام","الدين","الواحد","الكريم","اسحاق","موسي","هارون","ادم"}
def distinctive(toks): return {t for t in toks if t not in COMMON}
def father(nn):
    t=nn.split()
    for i,x in enumerate(t):
        if x in ("بن","ابن") and i+1<len(t):
            f=[t[i+1]]
            if f[0] in ("ابي","ابو","ابا","ام") and i+2<len(t): f.append(t[i+2])
            return tuple(f)
    return None
def father_compat(a,b):
    # same person => same father; tolerant of truncation & grandfather-attribution,
    # but a father made only of COMMON tokens (e.g. عبد الله) must not bridge.
    fa,fb=a["fath"],b["fath"]
    if not fa or not fb: return True
    if fa==fb: return True
    if set(fa)<=b["toks"] and (set(fa)-COMMON): return True
    if set(fb)<=a["toks"] and (set(fb)-COMMON): return True
    return False
def core_key(nn):
    t=content_toks(nn)
    return " ".join(t[:3]) if len(t)>=3 else " ".join(t)

# ---- Arabic death-year parser (reuse) ----
_U={"احدي":1,"واحده":1,"واحد":1,"اثنتين":2,"اثنين":2,"ثنتين":2,"ثلاث":3,"ثلاثه":3,"اربع":4,"اربعه":4,
    "خمس":5,"خمسه":5,"ست":6,"سته":6,"سبع":7,"سبعه":7,"ثمان":8,"ثماني":8,"ثمانيه":8,"تسع":9,"تسعه":9}
_T={"عشرين":20,"ثلاثين":30,"اربعين":40,"خمسين":50,"ستين":60,"سبعين":70,"ثمانين":80,"تسعين":90,"عشره":10,"عشر":10}
_H={"مايه":100,"ميه":100,"مايتين":200,"مئتين":200,"مايتي":200,"ثلاثمايه":300,"اربعمايه":400,"خمسمايه":500,
    "ستمايه":600,"سبعمايه":700,"ثمانمايه":800,"تسعمايه":900}
YEAR_AT=re.compile(r"(?:مات|توفي|توفيت|ماتت)\s+(?:سنه\s+)?")
def _base(w):
    if w in _H or w in _T or w in _U: return w
    if w.startswith("و") and (w[1:] in _H or w[1:] in _T or w[1:] in _U): return w[1:]
    return w
def parse_year(t):
    for m in YEAR_AT.finditer(t):
        toks=t[m.end():].split()[:8]; tot=0; i=0; got=False
        while i<len(toks):
            w=_base(toks[i])
            if w in _H: tot+=_H[w]; got=True
            elif w in _T: tot+=_T[w]; got=True
            elif w in _U:
                nb=_base(toks[i+1]) if i+1<len(toks) else ""
                if nb in ("عشره","عشر"): tot+=_U[w]+10; i+=1
                else: tot+=_U[w]
                got=True
            elif toks[i]=="و": pass
            else: break
            i+=1
        if got and 1<=tot<=1000: return tot
    return None

# ---- isnad neighbours: teachers (روى عن …) and students (وعنه …) from entry text ----
NB_STOP={"ابيه","امه","جده","اخيه","عمه","جماعه","خلق","ابوه","ابنه","عده","جمع","اخرين",
         "غيرهم","وغيره","وغيرهم","غير","واحد","ايضا","بنته","النبي","رسول الله","ابيها","اهل"}
def nbkey(piece):
    toks=[t for t in piece.split() if t not in ("بن","ابن","ابو","ابي","ابا")]
    return " ".join(toks[:3]).strip()
def splitnames(span):
    out=set()
    for piece in re.split(r"\s+و\s+|،|\s+ثم\s+", span[:90]):
        k=nbkey(piece)
        if len(k)>=4 and k not in NB_STOP: out.add(k)
        if len(out)>=6: break
    return out
STU=re.compile(r"(?:وعنه|روي عنه|حدث عنه|رواه عنه|يروي عنه|اخذ عنه)\s+(.+?)(?:\.|قال|مات|توفي|وثقه|ضعفه|$)")
TEA=re.compile(r"(?:روي عن(?= )|حدث عن(?= )|يروي عن(?= )|سمع من|اخذ عن(?= )|سمع)\s+(.+?)(?:\.|وعنه|روي عنه|قال|مات|توفي|$)")
def neighbours(t):
    tset=set(); sset=set()
    for m in STU.finditer(t): sset|=splitnames(m.group(1))
    for m in TEA.finditer(t): tset|=splitnames(m.group(1))
    return tset, sset

# ---- canonical authority: merged rijal DB (app/data/rijal — integrates GK/Jawāmiʿ
#      al-Kalim + classical sources; 115k profiles, ~37k numeric deaths, tabaqāt). ----
import json, glob
def _death(s):
    m=re.search(r"\d+", s or ""); return int(m.group()) if m else None
key2ids=defaultdict(set); id_death={}; id_name={}; id_tab={}
A_core=defaultdict(list)               # core_key -> [(tokset, id, death)] for fuzzy linking
short_era=defaultdict(list)            # short/peer name -> deaths (contemporary time-fix)
def _add_short(k,d):
    if k and d: short_era[k].append(d)
for f in glob.glob(os.path.join(ROOT,"app/data/rijal/profiles_*.json")):
    for pid,p in json.load(open(f,encoding="utf-8")).items():
        d=_death(p.get("death",""))
        id_death[pid]=d; id_name[pid]=p.get("full_name","")
        id_tab[pid]=p.get("tabaqat","").strip() if p.get("tabaqat","-") not in ("-","",None) else ""
        fn=norm(p.get("full_name",""))
        if not fn: continue
        fnism=(content_toks(fn) or [""])[0]
        # index full_name + only those namings that start with the person's ism
        # (drops relational/nasab-first forms like "أبيه"/father's name that conflate kin)
        forms={fn}|{norm(v) for v in p.get("namings",[]) if v and content_toks(norm(v)) and content_toks(norm(v))[0]==fnism}
        for v in forms:
            ct=content_toks(v); ts=set(ct)
            if len(ts)>=3:
                key2ids[v].add(pid); A_core[core_key(v)].append((ts,pid,d,ct[0]))
            t=[x for x in v.split() if x not in ("بن","ابن")]
            _add_short(" ".join(t[:3]),d); _add_short(" ".join(t[:2]),d)
            if t: _add_short(t[-1],d)

def link_canon(nn, toks, dyr):
    ids=key2ids.get(nn)
    if ids and len(ids)==1: return next(iter(ids))        # exact unique (safe any length)
    ct=content_toks(nn); nnism=ct[0] if ct else ""
    if len(toks)>=4:                                      # fuzzy only for specific-enough names
        best=None; bestj=0.0
        for ts,i,vd,vism in A_core.get(core_key(nn), ()):
            if vism!=nnism: continue                      # same ism (first name) — blocks father/son
            if vd and dyr and abs(vd-dyr)>5: continue
            uni=len(toks|ts)
            if not uni: continue
            j=len(toks&ts)/uni
            deathm=bool(vd and dyr and abs(vd-dyr)<=2)
            if (j>=0.80 or (j>=0.6 and deathm)) and j>bestj: best,bestj=i,j
        if best is not None: return best
    if ids and dyr is not None:                           # exact-ambiguous -> death disambiguation
        cand=[i for i in ids if id_death.get(i) and abs(id_death[i]-dyr)<=3]
        if len(cand)==1: return cand[0]
    return None

# ---- load entries ----
E=[]
for r in csv.DictReader(open(os.path.join(OUT,"unified_narrator_index.csv"),encoding="utf-8-sig")):
    nn=r["name_norm"]; toks=content_toks(nn)
    if len(toks)<2: continue
    tnorm=norm(r["text"]); dyr=parse_year(tnorm); tset,sset=neighbours(tnorm)
    E.append({"row":int(r["row_id"]),"slug":r["source_slug"],"book":r["source_book"],
              "page":r["page"],"name":r["narrator_name"],"nn":nn,"toks":set(toks),
              "ntok":len(toks),"dyr":dyr,"tset":tset,"sset":sset,"fath":father(nn),
              "aid":link_canon(nn,set(toks),dyr),"core":core_key(nn)})
print("entries:",len(E),"| with death yr:",sum(1 for e in E if e["dyr"]),
      "| with isnad nbrs:",sum(1 for e in E if e["tset"] or e["sset"]),
      "| canon-linked:",sum(1 for e in E if e["aid"] is not None))

# ---- TEMPORAL FIX: era estimate per entry (death > tabaqa > peers) ----
import statistics
tabmap={}                                  # name_norm -> tabaqa bin (from build_tabaqa output)
try:
    for r in csv.DictReader(open(os.path.join(OUT,"unified_by_tabaqa.csv"),encoding="utf-8-sig")):
        if r["tabaqa_order"]!="8": tabmap[r["name_norm"]]=int(r["tabaqa_order"])
except Exception as ex: print("  (tabaqa map unavailable:",ex,")")
TAB_MID={1:55,2:105,3:128,4:162,5:195,6:225,7:262}   # bin -> approx death-year midpoint

def _res(k):
    ds=short_era.get(k)
    if ds and max(ds)-min(ds)<=60: return statistics.median(ds)
    return None
def peer_era(e):
    ds=[]
    for k in (e["tset"]|e["sset"]):
        v=_res(k) or _res(k.split()[-1] if k.split() else "")
        if v: ds.append(v)
    return statistics.median(ds) if len(ds)>=2 else None
def era(e):
    if e["dyr"]: return (e["dyr"],4,"death")
    if e["aid"] is not None and id_death.get(e["aid"]): return (id_death[e["aid"]],6,"canon")
    t=tabmap.get(e["nn"])
    if t in TAB_MID: return (TAB_MID[t],22,"tabaqa")
    pe=peer_era(e)
    if pe is not None: return (pe,30,"peer")
    return None
for e in E: e["era"]=era(e)
print("  era fixed:",sum(1 for e in E if e["era"]),
      "(death",sum(1 for e in E if e["era"] and e["era"][2]=="death"),
      "| canon",sum(1 for e in E if e["era"] and e["era"][2]=="canon"),
      "| tabaqa",sum(1 for e in E if e["era"] and e["era"][2]=="tabaqa"),
      "| peer",sum(1 for e in E if e["era"] and e["era"][2]=="peer"),")")

# ---- union-find ----
parent=list(range(len(E)))
def find(x):
    while parent[x]!=x: parent[x]=parent[parent[x]]; x=parent[x]
    return x
def union(a,b):
    ra,rb=find(a),find(b)
    if ra!=rb: parent[rb]=ra

def compatible(a,b):
    # hard splits: different canonical ids, or conflicting death years
    if a["aid"] and b["aid"]:
        return ("aid", True) if a["aid"]==b["aid"] else (None, False)
    if a["dyr"] and b["dyr"] and abs(a["dyr"]-b["dyr"])>5: return (None, False)
    # TEMPORAL split: same name but incompatible era => different people
    ea,eb=a["era"],b["era"]
    era_ok=False
    if ea and eb:
        if abs(ea[0]-eb[0]) > ea[1]+eb[1]+20: return (None, False)
        era_ok = abs(ea[0]-eb[0]) <= ea[1]+eb[1]
    if not father_compat(a,b): return (None, False)        # same person => same father
    inter=len(a["toks"]&b["toks"]); uni=len(a["toks"]|b["toks"])
    if not uni: return (None, False)
    jac=inter/uni; mn=min(a["ntok"],b["ntok"])
    death_eq = bool(a["dyr"] and b["dyr"] and abs(a["dyr"]-b["dyr"])<=1)
    confident_era = era_ok and (a["era"][2] in ("death","canon","tabaqa") or b["era"][2] in ("death","canon","tabaqa"))
    sd=len(distinctive(a["toks"]) & distinctive(b["toks"]))   # shared *identifying* tokens
    # Jaccard (not containment); require shared distinctive token(s) so common names can't hub.
    if sd>=1 and mn>=4 and jac>=0.85: return ("name≈", True)
    if sd>=2 and mn>=4 and jac>=0.62 and death_eq: return ("name+death", True)
    # name + temporal fix (death/tabaqa/peer era agree): safe lower-Jaccard merge
    if sd>=2 and mn>=4 and jac>=0.55 and confident_era: return ("name+era", True)
    # isnad corroboration: same name-core block + share transmitters/students
    shared=len(a["tset"]&b["tset"])+len(a["sset"]&b["sset"])
    if sd>=1 and shared>=3 and jac>=0.45: return ("isnad", True)
    if sd>=1 and shared>=2 and jac>=0.6:  return ("isnad", True)
    return (None, False)

# index entries by canon id for cross-block merges
by_aid=defaultdict(list)
for i,e in enumerate(E):
    if e["aid"] is not None: by_aid[e["aid"]].append(i)
ev=Counter()
for ids in by_aid.values():                      # same canonical id, but require father + distinctive
    for x in range(len(ids)):
        ax=E[ids[x]]
        for y in range(x+1,len(ids)):
            bx=E[ids[y]]
            if father_compat(ax,bx) and (distinctive(ax["toks"]) & distinctive(bx["toks"])):
                if find(ids[x])!=find(ids[y]): union(ids[x],ids[y]); ev["aid"]+=1

# within-block merges
blocks=defaultdict(list)
for i,e in enumerate(E): blocks[e["core"]].append(i)
BIG=600
for core,idx in blocks.items():
    n=len(idx)
    if n<2: continue
    if n>BIG:   # only strong (near-equal name) merges in huge blocks
        for a in range(n):
            ea=E[idx[a]]
            for b in range(a+1,n):
                tag,ok=compatible(ea,E[idx[b]])
                if ok and tag!="aid": union(idx[a],idx[b]); ev[tag+"(big)"]+=1
        continue
    for a in range(n):
        for b in range(a+1,n):
            tag,ok=compatible(E[idx[a]],E[idx[b]])
            if ok and tag!="aid": union(idx[a],idx[b]); ev[tag]+=1

# NOTE: a cross-block isnad merge (pairs sharing rare transmitters across different
# name-blocks) was prototyped and REJECTED — it merged distinct people who share an
# isnad circle (e.g. the brothers al-Hasan & Ali b. Salih b. Hayy). Relatives/peers
# share teachers, so without a matching ism+father anchor it breaks precision.
# Isnad corroboration is therefore applied only WITHIN a name-core block (see compatible()).

# ---- build clusters ----
cl=defaultdict(list)
for i in range(len(E)): cl[find(i)].append(i)
# safety net: re-split each union with STRICT edges (father agreement AND a shared
# distinctive token) so no patronymic-conflict cluster can ship, however it formed.
groups=[]
for members in cl.values():
    if len(members)<=1: groups.append(members); continue
    n=len(members); p=list(range(n))
    def _f(x,p=p):
        while p[x]!=x: p[x]=p[p[x]]; x=p[x]
        return x
    for x in range(n):
        ax=E[members[x]]
        for y in range(x+1,n):
            bx=E[members[y]]
            if father_compat(ax,bx) and (distinctive(ax["toks"]) & distinctive(bx["toks"])):
                p[_f(x)]=_f(y)
    sub=defaultdict(list)
    for i in range(n): sub[_f(i)].append(members[i])
    groups.extend(sub.values())
row2cl={i:gid for gid,members in enumerate(groups) for i in members}
clusters=[]
for root,members in enumerate(groups):
    ms=[E[i] for i in members]
    aids={m["aid"] for m in ms if m["aid"] is not None}
    aid=next(iter(aids)) if len(aids)==1 else ""
    deaths=[m["dyr"] for m in ms if m["dyr"]]
    canon=max(ms,key=lambda m:m["ntok"])["name"]
    books=sorted({m["slug"] for m in ms})
    vnames=[]
    for m in ms:
        if m["name"] not in vnames: vnames.append(m["name"])
    clusters.append({"cluster_id":root,"canonical_name":canon,"canon_id":aid,
        "basis": ("canon" if aid!="" else "name"),
        "death": (Counter(deaths).most_common(1)[0][0] if deaths else (id_death.get(aid) or "")),
        "tabaqa": id_tab.get(aid,"") if aid!="" else "",
        "n_entries":len(ms),"n_books":len(books),"books":"; ".join(books),
        "variant_names":" ⟂ ".join(vnames),
        "members":" | ".join(f'{m["slug"]}:{m["page"]}' for m in ms)})
clusters.sort(key=lambda c:(-c["n_books"],-c["n_entries"]))

# ---- outputs ----
CC=["cluster_id","canonical_name","canon_id","basis","death","tabaqa","n_entries","n_books","books","variant_names","members"]
with open(os.path.join(OUT,"narrator_clusters.csv"),"w",encoding="utf-8-sig",newline="") as f:
    w=csv.writer(f); w.writerow(CC)
    for c in clusters: w.writerow([c[k] for k in CC])

# reviewable xlsx: only cross-book clusters (the verified duplicates)
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
wb=Workbook(); ws=wb.active; ws.title="duplicate_clusters"; ws.sheet_view.rightToLeft=True
hf=PatternFill("solid",fgColor="1F4E78"); hfont=Font(bold=True,color="FFFFFF")
arab=Alignment(horizontal="right",vertical="top",wrap_text=True,readingOrder=2); ctr=Alignment(horizontal="center",vertical="top")
cols=[("الاسم المعتمد",40),("canon_id",10),("الأساس",9),("الوفاة",8),("الطبقة",14),("# كتب",7),("# مواضع",8),("الكتب",34),("الأسماء المدمجة",70)]
ws.append([c[0] for c in cols])
for i,(t,wd) in enumerate(cols,1):
    c=ws.cell(1,i); c.fill=hf; c.font=hfont; c.alignment=Alignment(horizontal="center",wrap_text=True); ws.column_dimensions[get_column_letter(i)].width=wd
ws.freeze_panes="A2"
for c in [x for x in clusters if x["n_books"]>1]:
    ws.append([c["canonical_name"],c["canon_id"],c["basis"],c["death"],c["tabaqa"],c["n_entries"],c["n_books"],c["books"],c["variant_names"][:600]])
    rr=ws.max_row
    for col in (1,5,8,9): ws.cell(rr,col).alignment=arab
    for col in (2,3,4,6,7): ws.cell(rr,col).alignment=ctr
wb.save(os.path.join(OUT,"duplicate_clusters.xlsx"))
with open(os.path.join(OUT,"entry_to_cluster.csv"),"w",encoding="utf-8-sig",newline="") as f:
    w=csv.writer(f); w.writerow(["row_id","source_slug","page","name","cluster_id","canon_id"])
    for i,e in enumerate(E): w.writerow([e["row"],e["slug"],e["page"],e["name"],row2cl[i],e["aid"] if e["aid"] is not None else ""])

merged=[c for c in clusters if c["n_entries"]>1]
multibook=[c for c in clusters if c["n_books"]>1]
print(f"\nclusters: {len(clusters)}  (from {len(E)} entries; -{100*(1-len(clusters)/len(E)):.0f}% )")
print(f"  merged clusters (>1 entry): {len(merged)}")
print(f"  cross-book clusters (>1 book): {len(multibook)}")
print(f"  canon-anchored clusters  : {sum(1 for c in clusters if c['canon_id']!='')}")
print("merge evidence:", dict(ev))
print("\nsample cross-book clusters (deduped across books):")
for c in multibook[:10]:
    print(f"  [{c['n_books']}bk/{c['n_entries']}e d.{c['death']}] {c['canonical_name'][:40]}  <{c['books']}>")
