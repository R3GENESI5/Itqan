#!/usr/bin/env python3
"""Build a unified narrator index across the per-narrator OpenITI books."""
import os, re, csv, sys, io, glob
sys.path.insert(0, "/tmp/build")
from batch import parse, records_entry_mode          # reuse the exact parser
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

SRC = "/home/user/hadith/sources"
OUT = os.path.join(SRC, "unified"); os.makedirs(OUT, exist_ok=True)

# slug, display title, author, death_AH, source_type
BOOKS = [
 ("kawakib_nayyirat","al-Kawākib al-Nayyirāt","Ibn al-Kayyāl",939,"ikhtilāṭ"),
 ("lisan_mizan_ibnhajar","Lisān al-Mīzān","Ibn Ḥajar",852,"general/jarḥ+taʿdīl"),
 ("tarikh_kabir_bukhari","al-Tārīkh al-Kabīr","al-Bukhārī",256,"general"),
 ("kashif_dhahabi","al-Kāshif","al-Dhahabī",748,"general (six-books)"),
 ("mughni_duafa_dhahabi","al-Mughnī fī al-Ḍuʿafāʾ","al-Dhahabī",748,"jarḥ (ḍuʿafāʾ)"),
 ("siyar_dhahabi","Siyar Aʿlām al-Nubalāʾ","al-Dhahabī",748,"biography"),
 ("tabaqat_ibnsad","al-Ṭabaqāt al-Kubrā","Ibn Saʿd",230,"biography/ṭabaqāt"),
 ("duafa_ibnjawzi","al-Ḍuʿafāʾ wa-l-matrūkīn","Ibn al-Jawzī",597,"jarḥ (ḍuʿafāʾ)"),
 ("marifat_thiqat_ijli","Maʿrifat al-Thiqāt","al-ʿIjlī",261,"taʿdīl (thiqāt)"),
 ("duafa_uqayli","al-Ḍuʿafāʾ al-Kabīr","al-ʿUqaylī",322,"jarḥ (ḍuʿafāʾ)"),
 ("thiqat_ibnshahin","Tārīkh asmāʾ al-Thiqāt","Ibn Shāhīn",385,"taʿdīl (thiqāt)"),
 ("majruhin_ibnhibban","al-Majrūḥīn","Ibn Ḥibbān",354,"jarḥ (majrūḥīn)"),
 ("jami_tahsil_alai","Jāmiʿ al-Taḥṣīl","al-ʿAlāʾī",761,"marāsīl/connection"),
 ("duafa_nasai","al-Ḍuʿafāʾ wa-l-matrūkīn","al-Nasāʾī",303,"jarḥ (ḍuʿafāʾ)"),
 ("tabaqat_mudallisin_ibnhajar","Ṭabaqāt al-Mudallisīn","Ibn Ḥajar",852,"tadlīs"),
 ("ightibat_sibt_ibnajami","al-Ightibāṭ","Sibṭ Ibn al-ʿAjamī",841,"ikhtilāṭ"),
]

DIAC = re.compile(r"[ؐ-ًؚ-ٰٟۖ-ۭـ]")
def norm(s):
    s = DIAC.sub("", s or "")
    s = (s.replace("أ","ا").replace("إ","ا").replace("آ","ا").replace("ٱ","ا")
           .replace("ى","ي").replace("ئ","ي").replace("ؤ","و")
           .replace("ة","ه").replace("ء",""))
    s = re.sub(r"[^؀-ۿ\s]"," ", s)
    return re.sub(r"\s+"," ", s).strip()
def head(nrm, n=5):
    return " ".join(nrm.split()[:n])

LEAD_NUM   = re.compile(r"^[•\*\s]*\(?\s*\d+\s*\)?\s*[-–—.]*\s*")
LEAD_PAREN = re.compile(r"^\([^)]{1,10}\)\s*:?\s*")
TRAIL_RUM  = re.compile(r"\s*\*?\s*\([^)]*\)\s*\.?\s*$")
HONOR      = re.compile(r"\s*[،,]?\s*(رضي الله عنه[ماهم]*|رحمه الله|عليه السلام|رحمها الله)\s*\.?$")

def clean_name(h):
    s = (h or "").strip()
    s = LEAD_NUM.sub("", s); s = LEAD_PAREN.sub("", s); s = LEAD_NUM.sub("", s)
    for _ in range(3):
        s2 = HONOR.sub("", TRAIL_RUM.sub("", s))
        if s2 == s: break
        s = s2
    return s.strip(" .،,:-–—*")

rows = []
per_book = {}
for slug,title,author,death,stype in BOOKS:
    cand = glob.glob(os.path.join(SRC, slug, "*.mARkdown"))
    if not cand:
        print("  [skip] no mARkdown for", slug); continue
    path = cand[0]
    meta, logical = parse(path)
    recs = records_entry_mode(logical)
    per_book[slug] = len(recs)
    for r in recs:
        raw = r["name"].strip()
        nm = clean_name(raw)
        quality = "heading"
        if len(norm(nm).split()) < 2:                 # number-only/empty heading (e.g. Kashif)
            fb = " ".join(r["text"].split()[:8])
            if len(norm(fb).split()) >= 2:
                nm, quality = fb, "body_fallback"
            else:
                quality = "degenerate"
        nn = norm(nm)
        keyable = quality != "degenerate" and len(nn.split()) >= 3
        rows.append({
            "source_slug": slug, "source_book": title, "author": author,
            "death_ah": death, "source_type": stype, "section": r.get("section",""),
            "entry_no": r.get("num",""), "narrator_name": nm, "raw_heading": raw,
            "name_norm": nn, "name_head": head(nn), "name_quality": quality,
            "keyable": keyable,
            "page": (r["pages"][0] if r["pages"] else ""),
            "n_chars": len(r["text"]), "text": r["text"],
        })

# ---- long-form CSV (the master) ----
COLS = ["row_id","source_slug","source_book","author","death_ah","source_type",
        "section","entry_no","narrator_name","raw_heading","name_norm","name_head",
        "name_quality","keyable","page","n_chars","text"]
with io.open(os.path.join(OUT,"unified_narrator_index.csv"),"w",encoding="utf-8-sig",newline="") as f:
    w = csv.writer(f); w.writerow(COLS)
    for i,r in enumerate(rows,1):
        w.writerow([i]+[r[c] for c in COLS[1:]])

# ---- grouped by exact normalized name ----
from collections import defaultdict, Counter
g = defaultdict(list)
for r in rows:
    if r["keyable"]: g[r["name_norm"]].append(r)
grouped = []
for k, items in g.items():
    books = sorted({it["source_slug"] for it in items})
    disp = max((it["narrator_name"] for it in items), key=len)
    grouped.append({
        "name_norm": k, "display_name": disp,
        "n_books": len(books), "n_entries": len(items),
        "books": "; ".join(books),
        "source_types": "; ".join(sorted({it["source_type"] for it in items})),
        "pages": " | ".join(f"{it['source_slug']}:{it['page']}" for it in items if it['page']),
    })
grouped.sort(key=lambda x:(-x["n_books"], -x["n_entries"]))
GCOLS = ["name_norm","display_name","n_books","n_entries","source_types","books","pages"]
with io.open(os.path.join(OUT,"unified_by_narrator.csv"),"w",encoding="utf-8-sig",newline="") as f:
    w = csv.writer(f); w.writerow(GCOLS)
    for r in grouped: w.writerow([r[c] for c in GCOLS])

# ---- browsable xlsx: by-narrator + stats ----
wb = Workbook(); ws = wb.active; ws.title = "by_narrator"; ws.sheet_view.rightToLeft = True
hf=PatternFill("solid",fgColor="1F4E78"); hfont=Font(bold=True,color="FFFFFF")
arab=Alignment(horizontal="right",vertical="top",wrap_text=True,readingOrder=2)
ctr=Alignment(horizontal="center",vertical="top")
cols=[("الاسم (normalized)",34),("الاسم الكامل",46),("# كتب",7),("# مواضع",8),("نوع المصدر",26),("الكتب",40),("الصفحات",40)]
ws.append([c[0] for c in cols])
for i,(t,wd) in enumerate(cols,1):
    c=ws.cell(1,i); c.fill=hf; c.font=hfont; c.alignment=Alignment(horizontal="center",wrap_text=True)
    ws.column_dimensions[get_column_letter(i)].width=wd
ws.freeze_panes="A2"
for r in grouped:
    ws.append([r["name_norm"],r["display_name"],r["n_books"],r["n_entries"],r["source_types"],r["books"],r["pages"][:600]])
    rr=ws.max_row
    for c in (1,2,5,6,7): ws.cell(rr,c).alignment=arab
    for c in (3,4): ws.cell(rr,c).alignment=ctr

ws2=wb.create_sheet("stats"); ws2.sheet_view.rightToLeft=True
ws2.column_dimensions["A"].width=34; ws2.column_dimensions["B"].width=14
multi=sum(1 for r in grouped if r["n_books"]>=2)
stat=[("books indexed",len(BOOKS)),("total narrator rows",len(rows)),
      ("unique normalized names",len(grouped)),("names in ≥2 books",multi),
      ("names in ≥3 books",sum(1 for r in grouped if r["n_books"]>=3))]
ws2.append(["metric","value"])
for k,v in stat: ws2.append([k,v])
ws2.append([]); ws2.append(["— rows per source book —",""])
for slug,title,*_ in BOOKS: ws2.append([title, per_book[slug]])
wb.save(os.path.join(OUT,"unified_narrator_index.xlsx"))

print(f"books: {len(BOOKS)}  rows: {len(rows)}  unique-names: {len(grouped)}  in≥2 books: {multi}")
print("top multi-source narrators:")
for r in grouped[:8]:
    print(f"  {r['n_books']}bk/{r['n_entries']}e  {r['display_name'][:46]}  [{r['books']}]")
