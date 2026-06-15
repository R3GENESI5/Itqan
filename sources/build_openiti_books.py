#!/usr/bin/env python3
"""Download a set of OpenITI rijal/hadith texts, produce clean .txt + .xlsx + README each."""
import os, re, io, sys, time, urllib.request, traceback
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

OUT = "/tmp/build/out"
os.makedirs(OUT, exist_ok=True)

# slug, pipeline_id, url
TARGETS = [
 ("tarikh_kabir_bukhari","tarikh_kabir","https://raw.githubusercontent.com/OpenITI/0275AH/master/data/0256Bukhari/0256Bukhari.TarikhKabir/0256Bukhari.TarikhKabir.Shamela0000956-ara1.completed"),
 ("majruhin_ibnhibban","majruhin","https://raw.githubusercontent.com/OpenITI/0375AH/master/data/0354IbnHibbanBusti/0354IbnHibbanBusti.Majruhin/0354IbnHibbanBusti.Majruhin.Shia003101Vols-ara1.completed"),
 ("duafa_uqayli","duafa_uqayli","https://raw.githubusercontent.com/OpenITI/0325AH/master/data/0322AbuJacfarCuqayli/0322AbuJacfarCuqayli.DucafaKabir/0322AbuJacfarCuqayli.DucafaKabir.Shamela0013041-ara1.completed"),
 ("marifat_thiqat_ijli","marifat_thiqat","https://raw.githubusercontent.com/OpenITI/0275AH/master/data/0261AbuHasanCijli/0261AbuHasanCijli.MacrifatThiqat/0261AbuHasanCijli.MacrifatThiqat.JK000497-ara1.mARkdown"),
 ("lisan_mizan_ibnhajar","lisan_mizan","https://raw.githubusercontent.com/OpenITI/0875AH/master/data/0852IbnHajarCasqalani/0852IbnHajarCasqalani.LisanMizan/0852IbnHajarCasqalani.LisanMizan.Shamela0036357-ara1.mARkdown"),
 ("kashif_dhahabi","kashif","https://raw.githubusercontent.com/OpenITI/0750AH/master/data/0748Dhahabi/0748Dhahabi.Kashif/0748Dhahabi.Kashif.Shia003276Vols-ara1.mARkdown"),
 ("mughni_duafa_dhahabi","mughni_duafa","https://raw.githubusercontent.com/OpenITI/0750AH/master/data/0748Dhahabi/0748Dhahabi.MughniFiDucafa/0748Dhahabi.MughniFiDucafa.JK001307-ara1.mARkdown"),
 ("siyar_dhahabi","siyar_alam_nubala","https://raw.githubusercontent.com/OpenITI/0750AH/master/data/0748Dhahabi/0748Dhahabi.SiyarAclamNubala/0748Dhahabi.SiyarAclamNubala.Shamela0010906-ara1.mARkdown"),
 ("duafa_nasai","duafa_nasai","https://raw.githubusercontent.com/OpenITI/0325AH/master/data/0303Nasai/0303Nasai.DucafaWaMatrukin/0303Nasai.DucafaWaMatrukin.JK000509-ara2.completed"),
 ("duafa_ibnjawzi","duafa_ibnjawzi","https://raw.githubusercontent.com/OpenITI/0600AH/master/data/0597IbnJawzi/0597IbnJawzi.DucafaWaMatrukin/0597IbnJawzi.DucafaWaMatrukin.Shamela0005830-ara1.completed"),
 ("tarikh_ibnmain","tarikh_ibnmain","https://raw.githubusercontent.com/OpenITI/0250AH/master/data/0233YahyaIbnMacin/0233YahyaIbnMacin.MacrifatRijal/0233YahyaIbnMacin.MacrifatRijal.Shamela0000101-ara1"),
 ("tabaqat_ibnsad","tabaqat_kubra","https://raw.githubusercontent.com/OpenITI/0250AH/master/data/0230IbnSacd/0230IbnSacd.TabaqatKubra/0230IbnSacd.TabaqatKubra.ShamAY0035884-ara1.mARkdown"),
 ("thiqat_ibnshahin","thiqat_ibnshahin","https://raw.githubusercontent.com/OpenITI/0400AH/master/data/0385IbnShahin/0385IbnShahin.TarikhAsmaThiqat/0385IbnShahin.TarikhAsmaThiqat.JK000511-ara1.completed"),
 ("tabaqat_mudallisin_ibnhajar","tabaqat_mudallisin","https://raw.githubusercontent.com/OpenITI/0875AH/master/data/0852IbnHajarCasqalani/0852IbnHajarCasqalani.TacrifAhlTaqdis/0852IbnHajarCasqalani.TacrifAhlTaqdis.Shia003340BK1-ara1.mARkdown"),
 ("marasil_ibnabihatim","marasil","https://raw.githubusercontent.com/OpenITI/0350AH/master/data/0327IbnAbiHatimRazi/0327IbnAbiHatimRazi.Marasil/0327IbnAbiHatimRazi.Marasil.JK000743-ara1"),
 ("jami_tahsil_alai","jami_tahsil","https://raw.githubusercontent.com/OpenITI/0775AH/master/data/0761IbnKaykaldiCalai/0761IbnKaykaldiCalai.JamicTahsil/0761IbnKaykaldiCalai.JamicTahsil.Shamela0025864-ara1.completed"),
 ("ightibat_sibt_ibnajami","ightibat","https://raw.githubusercontent.com/OpenITI/0850AH/master/data/0841BurhanDinSibtIbnCajami/0841BurhanDinSibtIbnCajami.Ightibat/0841BurhanDinSibtIbnCajami.Ightibat.Shamela0000130-ara1.completed"),
 ("ilal_daraqutni","ilal_daraqutni","https://raw.githubusercontent.com/OpenITI/0400AH/master/data/0385Daraqutni/0385Daraqutni.CilalWarida/0385Daraqutni.CilalWarida.Shamela0009082-ara1.completed"),
 ("ilal_ibnabihatim","ilal_ibnabihatim","https://raw.githubusercontent.com/OpenITI/0350AH/master/data/0327IbnAbiHatimRazi/0327IbnAbiHatimRazi.CilalHadith/0327IbnAbiHatimRazi.CilalHadith.JK000682-ara1"),
 ("muqaddimat_ibnsalah","muqaddimat_ibnsalah","https://raw.githubusercontent.com/OpenITI/0650AH/master/data/0643IbnSalahShahrazuri/0643IbnSalahShahrazuri.MuqaddimatCulumHadith/0643IbnSalahShahrazuri.MuqaddimatCulumHadith.JK000537-ara1.completed"),
 ("tadrib_rawi_suyuti","tadrib_rawi","https://raw.githubusercontent.com/OpenITI/0925AH/master/data/0911Suyuti/0911Suyuti.TadribRawi/0911Suyuti.TadribRawi.JK000138-ara1"),
 ("fath_mughith_sakhawi","fath_mughith","https://raw.githubusercontent.com/OpenITI/0925AH/master/data/0902Sakhawi/0902Sakhawi.FathMughith/0902Sakhawi.FathMughith.JK006675-ara1"),
 ("kifaya_khatib","kifaya_riwaya","https://raw.githubusercontent.com/OpenITI/0475AH/master/data/0463KhatibBaghdadi/0463KhatibBaghdadi.KifayaFiCilmRiwaya/0463KhatibBaghdadi.KifayaFiCilmRiwaya.JK000135-ara1"),
]

PAGE = re.compile(r"PageV(\d+)P(\d+)")
ENTRY = re.compile(r"^###\s*\$+")
SECT  = re.compile(r"^###\s*\|")
XLMAX = 32600

def fetch(url, dest, tries=4):
    if os.path.exists(dest) and os.path.getsize(dest) > 0:
        return True
    for i in range(tries):
        try:
            req = urllib.request.Request(url, headers={"User-Agent":"Mozilla/5.0 openiti-fetch"})
            with urllib.request.urlopen(req, timeout=120) as r:
                io.open(dest,"wb").write(r.read())
            return True
        except Exception as e:
            if i==tries-1:
                print("   FETCH FAIL:", e); return False
            time.sleep(2**i)

def parse(path):
    raw = io.open(path, encoding="utf-8", errors="replace").read().split("\n")
    try: start = next(i for i,l in enumerate(raw) if l.strip()=="#META#Header#End#")+1
    except StopIteration: start = 0
    meta = {}
    for l in raw[:start]:
        m = re.match(r"#META#\s*\d+\.(\w+)\s*::\s*(.*)", l)
        if m and m.group(2).strip() not in ("NODATA","NOTGIVEN",""):
            meta.setdefault(m.group(1), m.group(2).strip())
    logical = []
    for l in raw[start:]:
        if l.startswith("~~"):
            if logical: logical[-1]+=" "+l[2:]
            else: logical.append(l[2:])
        else: logical.append(l)
    return meta, logical

def pages_in(t):
    out=[]
    for v,p in PAGE.findall(t):
        v,p=int(v),int(p)
        if not(v==0 and p==0): out.append(f"ج{v}/ص{p}")
    return out
def clean_inline(t):
    t=PAGE.sub("",t); t=re.sub(r"(?<!\S)ms\d+(?!\S)","",t)
    return re.sub(r"[ \t]{2,}"," ",t).strip()

def to_clean_txt(meta, logical):
    out=[]
    for l in logical:
        s=l.strip()
        if not s or s.startswith("######OpenITI"): continue
        s=re.sub(r"(?<!\S)ms\d+(?!\S)","",s)
        s=PAGE.sub(lambda m:("" if (int(m.group(1))==0 and int(m.group(2))==0) else f"\n【ج{int(m.group(1))} ص{int(m.group(2))}】"), s)
        if ENTRY.match(s): out.append("\n\n### "+re.sub(r"^###\s*\$+\s*","",s).strip())
        elif SECT.match(s): out.append("\n\n## "+re.sub(r"^###\s*\|\s*","",s).strip())
        else:
            s=re.sub(r"^#{1,6}\s*","",s).strip()
            if s: out.append(s)
    txt="\n".join(out); txt=re.sub(r"\n{3,}","\n\n",txt).strip()+"\n"
    head=f"{meta.get('BookTITLE','')}\nالمؤلف: {meta.get('AuthorNAME','')}\nالمحقق: {meta.get('EdEDITOR','NODATA')}\nالناشر: {meta.get('EdPUBLISHER','NODATA')}\nالمصدر: OpenITI corpus\n"+"="*70+"\n\n"
    return head+txt

def records_entry_mode(logical):
    recs=[]; section=""; cur=None
    def flush():
        nonlocal cur
        if cur is not None:
            cur["pages"]=pages_in(cur["_raw"]); cur["text"]=clean_inline(cur["_raw"]); del cur["_raw"]; recs.append(cur)
            cur=None
    for l in logical:
        s=l.strip()
        if not s or s.startswith("######OpenITI"): continue
        if SECT.match(s):
            flush(); t=re.sub(r"^###\s*\|\s*","",s).strip()
            if not t.upper().startswith("APPENDIX"): section=t
            continue
        if ENTRY.match(s):
            flush(); h=re.sub(r"^###\s*\$+\s*","",s).strip()
            m=re.match(r"(\d+)\s*[-–]\s*(.*)",h)
            recs_num=m.group(1) if m else ""
            name=clean_inline(m.group(2) if m else h)
            cur={"section":section,"num":recs_num,"name":name,"_raw":""}; continue
        s=re.sub(r"^#{1,6}\s*","",s)
        if cur is not None: cur["_raw"]+=" "+s
    flush()
    return recs

def records_section_mode(logical):
    recs=[]; cur=None
    def flush():
        nonlocal cur
        if cur is not None:
            cur["pages"]=pages_in(cur["_raw"]); cur["text"]=clean_inline(cur["_raw"]); del cur["_raw"]; recs.append(cur)
            cur=None
    for l in logical:
        s=l.strip()
        if not s or s.startswith("######OpenITI"): continue
        if SECT.match(s) or ENTRY.match(s):
            flush(); h=re.sub(r"^###\s*[\|\$]+\s*","",s).strip()
            cur={"section":"","num":"","name":clean_inline(h),"_raw":""}; continue
        s=re.sub(r"^#{1,6}\s*","",s)
        if cur is None: cur={"section":"","num":"","name":"(صدر الكتاب)","_raw":""}
        cur["_raw"]+=" "+s
    flush()
    return recs

def records_page_mode(logical):
    recs=[]; buf=""
    for l in logical:
        s=l.strip()
        if not s or s.startswith("######OpenITI"): continue
        s=re.sub(r"^###\s*[\|\$]+\s*","",s); s=re.sub(r"^#{1,6}\s*","",s)
        for part in re.split(r"(PageV\d+P\d+)", s):
            m=re.fullmatch(r"PageV(\d+)P(\d+)", part or "")
            if m:
                v,p=int(m.group(1)),int(m.group(2))
                lbl="" if (v==0 and p==0) else f"ج{v}/ص{p}"
                t=clean_inline(buf)
                if t: recs.append({"section":"","num":"","name":"","pages":([lbl] if lbl else []),"text":t})
                buf=""
            else:
                buf+=" "+(part or "")
    t=clean_inline(buf)
    if t: recs.append({"section":"","num":"","name":"","pages":[],"text":t})
    return recs

def build_xlsx(meta, logical, dest, label):
    n_entry=sum(1 for l in logical if ENTRY.match(l.strip()))
    mode = "entry" if n_entry>=20 else "section"
    recs = records_entry_mode(logical) if mode=="entry" else records_section_mode(logical)
    if mode=="section" and len(recs)<10:
        pm=records_page_mode(logical)
        if len(pm)>len(recs): recs, mode = pm, "page"
    wb=Workbook(); ws=wb.active; ws.title=("التراجم" if mode=="entry" else "الأقسام")
    ws.sheet_view.rightToLeft=True
    arab=Alignment(horizontal="right",vertical="top",wrap_text=True,readingOrder=2)
    ctr =Alignment(horizontal="center",vertical="top")
    hf=PatternFill("solid",fgColor="1F4E78"); hfont=Font(bold=True,color="FFFFFF",size=12)
    cols=[("#",6),("القسم",22),("العنوان / الراوي",40),("الصفحة",11),("الصفحات",18),("النص",110),("أحرف",8)]
    ws.append([c[0] for c in cols])
    for i,(t,w) in enumerate(cols,1):
        c=ws.cell(1,i); c.fill=hf; c.font=hfont; c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width=w
    ws.freeze_panes="A2"
    for idx,r in enumerate(recs,1):
        txt=r["text"]
        if len(txt)>XLMAX: txt=txt[:XLMAX]+" …[مقتطع؛ النص الكامل في ملف clean.txt]"
        ws.append([idx,r["section"],r["name"][:300],(r["pages"][0] if r["pages"] else ""),"، ".join(r["pages"][:40]),txt,len(r["text"])])
        rr=ws.max_row
        ws.cell(rr,3).alignment=arab; ws.cell(rr,6).alignment=arab
        ws.cell(rr,1).alignment=ctr; ws.cell(rr,4).alignment=ctr; ws.cell(rr,7).alignment=ctr
        if rr%50==0 or len(recs)<400: ws.row_dimensions[rr].height=80
    # info sheet
    ws2=wb.create_sheet("معلومات"); ws2.sheet_view.rightToLeft=True
    ws2.column_dimensions["A"].width=24; ws2.column_dimensions["B"].width=95
    info=[("الكتاب",meta.get("BookTITLE",label)),("المؤلف",meta.get("AuthorNAME","")),
          ("المحقق",meta.get("EdEDITOR","NODATA")),("الناشر",meta.get("EdPUBLISHER","NODATA")),
          ("الطبعة",meta.get("EdNUMBER","NODATA")),("المصدر","OpenITI corpus"),
          ("بنية الجدول",{"entry":"سطر لكل ترجمة","section":"سطر لكل قسم/باب","page":"سطر لكل صفحة"}[mode]),
          ("عدد السطور",str(len(recs))),
          ("ملاحظة","النص منقول حرفيًا؛ روابط الصفحات بصيغة ج/ص؛ لا استنتاج آلي للأحكام.")]
    for i,(k,v) in enumerate(info,1):
        ws2.cell(i,1,k).font=Font(bold=True); ws2.cell(i,1).alignment=arab; ws2.cell(i,2,v).alignment=arab
    wb.save(dest)
    return mode, len(recs)

def readme(meta, slug, url, mode, nrec):
    return f"""# {meta.get('BookTITLE',slug)}

- **Author:** {meta.get('AuthorNAME','NODATA')}
- **Editor:** {meta.get('EdEDITOR','NODATA')}
- **Publisher / ed.:** {meta.get('EdPUBLISHER','NODATA')} — {meta.get('EdNUMBER','NODATA')}
- **Source:** OpenITI corpus (primary version), fetched from GitHub raw.

OpenITI version URL:
<{url}>

## Files
- `{slug}.mARkdown` — primary OpenITI text (structured; `### $` entries, `PageVxxPyyy` page markers, `#META#` header). Canonical for ingestion.
- `{slug}_clean.txt` — readable plain-text (soft-wraps joined, milestones stripped, pages as 【ج.ص】).
- `{slug}.xlsx` — structured workbook ({ {"entry":"one row per biographical entry","section":"one row per section/chapter","page":"one row per page"}[mode] }; {nrec} rows) + provenance sheet.

Text is reproduced verbatim from the primary edition; no automated grading inference.
"""

def main():
    summary=[]
    pipe=[]
    for slug,pid,url in TARGETS:
        d=os.path.join(OUT,slug); os.makedirs(d,exist_ok=True)
        ext = "."+url.rsplit("-ara",1)[1].split(".",1)[1] if "." in url.rsplit("-ara",1)[-1] else ""
        md=os.path.join(d,slug+(".mARkdown" if "mARkdown" in url else ".txt" if ext=="" else ".mARkdown" if False else ".mARkdown"))
        md=os.path.join(d,slug+".mARkdown")
        print(f"[{slug}] downloading…")
        if not fetch(url, md):
            summary.append((slug,"FETCH_FAIL",0,0)); continue
        size=os.path.getsize(md)
        try:
            meta,logical=parse(md)
            io.open(os.path.join(d,slug+"_clean.txt"),"w",encoding="utf-8").write(to_clean_txt(meta,logical))
            mode,nrec=build_xlsx(meta,logical,os.path.join(d,slug+".xlsx"),slug)
            io.open(os.path.join(d,"README.md"),"w",encoding="utf-8").write(readme(meta,slug,url,mode,nrec))
            title=meta.get("BookTITLE",slug)
            pipe.append((pid,title,url,pid+".txt"))
            summary.append((slug,mode,nrec,size))
            print(f"   ok  {mode:7} rows={nrec:5} src={size//1024}KB  {title}")
        except Exception as e:
            print("   BUILD FAIL:",e); traceback.print_exc(); summary.append((slug,"BUILD_FAIL",0,size))
    print("\n==== SUMMARY ====")
    for s in summary: print(f"  {s[0]:34} {s[1]:10} rows={s[2]:<6} {s[3]//1024}KB")
    # emit pipeline snippet
    with io.open(os.path.join(OUT,"_pipeline_entries.txt"),"w",encoding="utf-8") as f:
        for pid,title,url,fn in pipe:
            f.write('    {\n')
            f.write(f'        "id": "{pid}",\n')
            f.write(f'        "title": "{title}",\n')
            f.write(f'        "url": "{url}",\n')
            f.write(f'        "filename": "{fn}",\n')
            f.write('    },\n')
    print("\nwrote pipeline entries:", os.path.join(OUT,"_pipeline_entries.txt"))

if __name__=="__main__":
    main()
